# -*- coding: utf-8 -*-
"""生成 office-doc-grading skill 评测样例（标准文档 + 需求文本 + 学生文件）。"""
import sys
from pathlib import Path

sys.path.insert(0, r"E:\opencode\GradingServer\testbed")
from builder import build_word, build_excel, build_ppt  # noqa: E402
from docx.enum.text import WD_ALIGN_PARAGRAPH  # noqa: E402

OUT = Path(__file__).parent / "files"
OUT.mkdir(parents=True, exist_ok=True)


def body_paras(texts, size=12, font="宋体", align=None):
    return [{"text": t, "size": size, "font": font, "align": align} for t in texts]


def build_word_with_margins(margins, **kw):
    """build_word 后回填四边距（builder 只支持 top）。margins=(top, bottom, left, right) 厘米。"""
    from docx import Document
    from docx.shared import Cm
    from io import BytesIO

    data = build_word(**kw)
    doc = Document(BytesIO(data))
    sec = doc.sections[0]
    top, bottom, left, right = margins
    sec.top_margin = Cm(top)
    sec.bottom_margin = Cm(bottom)
    sec.left_margin = Cm(left)
    sec.right_margin = Cm(right)
    buf = BytesIO()
    doc.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# eval 1: word 论文排版
# ---------------------------------------------------------------------------
standard_texts = body_paras([
    "本文档为《计算机应用基础》课程论文排版标准。",
    "论文应包含标题、摘要、正文章节与参考文献。",
    "页边距要求：上 2.5 厘米、下 2.5 厘米、左 3.0 厘米、右 2.6 厘米。",
    "标题使用黑体三号（16 磅）居中加粗；正文使用宋体小四（12 磅），1.5 倍行距。",
    "页眉必须包含“课程论文”字样；页脚需有页码。",
    "正文字数不少于 200 字。",
])
(OUT / "标准_论文排版.docx").write_bytes(build_word_with_margins(
    (2.5, 2.5, 3.0, 2.6),
    title="课程论文排版标准",
    title_size=16, title_bold=True, title_align=WD_ALIGN_PARAGRAPH.CENTER,
    body=[dict(t, size=12) for t in standard_texts],
    margin_top=2.5,
    header_text="XX大学 课程论文",
    footer_text="第 1 页",
))
REQ_WORD = """请按以下要求评级《计算机应用基础》课程论文作业（word 文档）：
1. 页边距：上 2.5cm、下 2.5cm、左 3.0cm、右 2.6cm，允许 ±0.3cm 偏差；
2. 标题必须居中加粗；
3. 正文为宋体、小四（12 磅）；
4. 页眉必须包含“课程论文”；
5. 页脚必须有页码；
6. 论文字数不少于 200 字。
各项权重均匀分配，总分 100。"""
(OUT / "需求_论文排版.txt").write_text(REQ_WORD, encoding="utf-8")

# 合规学生
zhang_texts = body_paras([
    "1. 引言", "计算机应用基础课程旨在培养学生的信息化素养与实际操作能力。",
    "在信息化时代，办公软件的应用能力已成为大学生必备的基本技能之一。",
    "本文围绕文档排版、数据处理与演示制作三个模块展开论述。",
    "2. 文档排版", "文档排版是办公自动化中最基础也最重要的环节之一。",
    "合理的页面设置、规范的字体字号与行距，直接影响文档的可读性与美观度。",
    "Word 提供了丰富的排版工具，包括页边距、纸张方向、页眉页脚与页码设置。",
    "3. 数据处理", "Excel 强大的公式与函数功能，能够高效完成数据的统计与分析。",
    "条件格式与筛选功能让大量数据的呈现更加直观。",
    "4. 演示制作", "PowerPoint 通过幻灯片版式、动画与切换效果展现演示内容。",
    "5. 结语", "熟练掌握办公软件是当代大学生的基本素养，也是职业发展的基础。",
    "希望每一位同学都能在实践中不断提升自己的信息化应用能力。",
    "参考文献", "[1] 计算机应用基础教程. 高等教育出版社, 2023. [2] Office 办公自动化实用教程. 清华大学出版社, 2022.",
])
(OUT / "张三_论文.docx").write_bytes(build_word_with_margins(
    (2.5, 2.5, 3.0, 2.6),
    title="办公软件应用能力培养研究",
    title_size=16, title_bold=True, title_align=WD_ALIGN_PARAGRAPH.CENTER,
    body=zhang_texts,
    margin_top=2.5,
    header_text="XX大学 课程论文",
    footer_text="第 1 页",
))
# 违规学生：页边距错、标题不加粗、正文 14 磅、无页眉页脚、字数不够
li_texts = body_paras([
    "课程论文",
    "计算机应用基础是一门重要的课程。",
    "本文简单介绍办公软件。",
], size=14)
(OUT / "李四_论文.docx").write_bytes(build_word_with_margins(
    (3.5, 2.0, 3.0, 2.6),
    title="课程论文", title_size=16, title_bold=False,
    body=li_texts, margin_top=3.5,
))

# ---------------------------------------------------------------------------
# eval 2: xlsx 成绩表
# ---------------------------------------------------------------------------
def build_xlsx(**kw):
    """build_excel 后补 merge（builder 的 merge+table 组合会写 MergedCell 崩溃）。"""
    from openpyxl import load_workbook
    from io import BytesIO

    data = build_excel(**{k: v for k, v in kw.items() if k != "merge"})
    if kw.get("merge"):
        wb = load_workbook(BytesIO(data))
        ws = wb.active
        if ws["A1"].value is None:
            ws["A1"] = "成绩表"
        ws.merge_cells("A1:B1")
        buf = BytesIO()
        wb.save(buf)
        data = buf.getvalue()
    return data


(OUT / "标准_成绩表.xlsx").write_bytes(build_xlsx(
    formula=True, freeze=True, merge=True, table=True, conditional=True,
    auto_filter=True, border=True, font_size=11, sheets=2,
))
REQ_XLSX = """请按以下要求评分《数据分析》上机作业（excel 成绩表）：
1. 工作表中必须包含公式（如求和/平均分）；
2. 必须冻结窗格（表头行）；
3. 标题行需合并单元格；
4. 需使用表格样式；
5. 需设置筛选；
6. 需有条件格式（如不及格标红）；
7. 需有单元格边框；
8. 工作表数量 2 个以上。
各项权重均匀分配，总分 100。"""
(OUT / "需求_成绩表.txt").write_text(REQ_XLSX, encoding="utf-8")

(OUT / "王五_成绩表.xlsx").write_bytes(build_xlsx(
    formula=True, freeze=True, merge=True, table=True, conditional=True,
    auto_filter=True, border=True, font_size=11, sheets=2,
))
(OUT / "赵六_成绩表.xlsx").write_bytes(build_excel())  # 只有基础数据
(OUT / "孙七_成绩表.xlsx").write_bytes(build_xlsx(
    formula=True, freeze=True, sheets=1,  # 缺合并/表格/筛选/条件格式/边框/多表
))

# ---------------------------------------------------------------------------
# eval 3: pptx 演示文稿
# ---------------------------------------------------------------------------
(OUT / "标准_演示文稿.pptx").write_bytes(build_ppt(
    slides=6, font_size=24, notes=True, transition=True, animation=True,
    images=1, center=True,
))
REQ_PPTX = """请按以下要求评分《职业规划》演示文稿作业（ppt 文件）：
1. 幻灯片不少于 6 页；
2. 页面正文字号不小于 24 磅；
3. 每页需有演讲者备注；
4. 需包含至少 1 张图片；
5. 需有切换效果；
6. 需有动画效果；
7. 文本应居中。
各项权重均匀分配，总分 100。"""
(OUT / "需求_演示文稿.txt").write_text(REQ_PPTX, encoding="utf-8")

(OUT / "钱七_演示.pptx").write_bytes(build_ppt(
    slides=6, font_size=24, notes=True, transition=True, animation=True,
    images=1, center=True,
))
(OUT / "孙八_演示.pptx").write_bytes(build_ppt(
    slides=3, font_size=14,  # 缺备注/图片/切换/动画/居中
))

print("files ->", sorted(p.name for p in OUT.iterdir()))