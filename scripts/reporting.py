# -*- coding: utf-8 -*-
"""报告生成（P6）：summary.xlsx（每学生一份 sheet）+ reports/<学生名>_report.html。

技能规定：评分结果必须是可审计的，每一条都要有"实际值 vs 期望 + 位置依据"。
本模块把报告生成固化为确定性脚本，避免每次执行者手搓、结构漂移。

输入：results（list[dict]），每条：
    {criterion: {label, weight, deduct, scope}, ok: bool, weight: float,
     evidence: str, resolved: bool, need_review: bool, actual: any,
     expected: any, comparator: str, position: str}
输出：
    summary.xlsx           —— 每名学生一个工作表(sheet)，逐条明细（评分项/结论/分值/
                              实际值/期望值/位置/依据/扣分说明），与 HTML 报告明细同构
    reports/<name>_report.html —— 逐条表（相同列，供快速查看）

用法（每份学生作业调用一次，自动向同一 workbook 追加 sheet）：
    from reporting import build_report
    build_report(student_name, file_name, results, scoring, out_dir)
"""
from __future__ import annotations

import re
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except Exception:  # pragma: no cover
    _HAS_OPENPYXL = False

LEVELS = ((90, "优秀"), (75, "良好"), (60, "合格"), (0, "需改进"))
SUMMARY_FILE = "summary.xlsx"

# 与 HTML 报告一致的表头
HEADERS = ["评分项", "结论", "分值", "实际值", "期望值", "位置", "依据", "扣分说明"]
COL_WIDTHS = [24, 8, 8, 22, 22, 18, 40, 24]

_FILL_PASS = PatternFill("solid", start_color="E8F5E9")
_FILL_FAIL = PatternFill("solid", start_color="FFEBEE")
_FILL_REVIEW = PatternFill("solid", start_color="FFF8E1")
_FILL_HEAD = PatternFill("solid", start_color="F5F5F5")
_FONT_HEAD = Font(bold=True)
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
_TOP = Alignment(horizontal="left", vertical="top")


def level_of(pct):
    for th, name in LEVELS:
        if pct >= th:
            return name
    return "需改进"


def _fmt(v):
    if v is None:
        return "—"
    if isinstance(v, (set, list, tuple, dict)):
        import json
        try:
            return json.dumps(v, ensure_ascii=False)
        except (TypeError, ValueError):
            return str(v)
    return str(v)


def _position_from_scope(scope):
    if not scope:
        return ""
    t = scope.get("type")
    if t == "title":
        return "标题段落"
    if t == "body":
        return "正文章节"
    if t == "text":
        return "含“%s”的段落" % scope.get("val", "")
    if t == "para":
        return "第 %s 段" % scope.get("index")
    if t == "table":
        return "第 %s 个表格" % scope.get("index")
    if t == "image":
        return "第 %s 张图片" % scope.get("index")
    if t == "header":
        return "页眉"
    if t == "footer":
        return "页脚"
    return str(scope)


def _sheet_name(student_name, wb):
    """由学生名生成合法 sheet 名（<=31 字符，去非法字符）。同名 sheet 以"替换"刷新，不追加序号。"""
    name = re.sub(r'[:\\/?*\[\]]', "_", str(student_name)).strip() or "学生"
    if len(name.encode("utf-8", "ignore")) > 31:
        name = name[:31]
    return name or "学生"


def _load_or_create(out_dir):
    """打开既有汇总 workbook（追加），否则新建。返回 (wb, path)。"""
    path = Path(out_dir) / SUMMARY_FILE
    if path.exists() and _HAS_OPENPYXL:
        try:
            return load_workbook(path), path
        except Exception:
            pass
    return Workbook(), path


def _write_student_sheet(wb, student_name, file_name, results, scoring, uncovered=None):
    """向 wb 写入一个学生的明细 sheet（与 HTML 报告同构）。"""
    pct = scoring["pct"]
    level = scoring.get("level") or level_of(pct)
    title = _sheet_name(student_name, wb)
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title=title)

    try:
        ws.sheet_view.showGridLines = False
    except Exception:
        pass

    # ---- 信息区 ----
    ws["A1"] = "作业评分报告"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "文件：%s" % file_name
    ws["A3"] = "学生：%s" % student_name
    ws["A4"] = "总分 %s / %s（%s%%）　等级：%s" % (
        scoring["score"], scoring["max_score"], pct, level)
    try:
        ws["A4"].font = Font(size=12, color="1565C0", bold=True)
    except Exception:
        pass
    info = {
        "通过项": sum(1 for r in results if r["ok"]),
        "待复核项": sum(1 for r in results if r.get("need_review")),
        "未覆盖需求": "；".join(uncovered) if uncovered else "无",
    }
    r = 5
    for k, v in info.items():
        ws.cell(r, 1, "%s：%s" % (k, v))
        r += 1

    # ---- 明细表头 ----
    head_row = r + 1
    for ci, h in enumerate(HEADERS, start=1):
        cell = ws.cell(head_row, ci, h)
        cell.font = _FONT_HEAD
        cell.fill = _FILL_HEAD
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    ws.row_dimensions[head_row].height = 20

    # ---- 明细行 ----
    rw = head_row + 1
    for res in results:
        c = res["criterion"]
        verdict = "通过" if res["ok"] else ("待复核" if res.get("need_review") else "未通过")
        pos = res.get("position") or _position_from_scope(c.get("scope"))
        deduct = c.get("deduct", "") or ("" if res["ok"] else "未达到要求")
        values = [
            c.get("label", ""), verdict, res["weight"],
            _fmt(res.get("actual")), _fmt(res.get("expected")),
            pos, res.get("evidence", ""), deduct,
        ]
        fill = _FILL_PASS if res["ok"] else (_FILL_REVIEW if res.get("need_review") else _FILL_FAIL)
        for ci, v in enumerate(values, start=1):
            cell = ws.cell(rw, ci, v)
            cell.border = _BORDER
            cell.alignment = _WRAP if ci in (4, 5, 7, 8) else _TOP
            if ci == 2:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            if ci == 3:
                cell.alignment = Alignment(horizontal="right", vertical="top")
            cell.fill = fill
        rw += 1

    # ---- 列宽/冻结 ----
    for ci, wth in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = wth
    try:
        ws.freeze_panes = "A%d" % (head_row + 1)
    except Exception:
        pass
    return ws


def build_report(student_name, file_name, results, scoring, out_dir, uncovered=None):
    """scoring: {score, max_score, pct, level}（可由 score_diffs 得到，也可传入内部分数）。"""
    out = Path(out_dir)

    # ---- XLSX：每学生一个 sheet，追加到共享 workbook ----
    xlsx_path = None
    if _HAS_OPENPYXL:
        wb, xlsx_path = _load_or_create(out)
        for _sn in list(wb.sheetnames):
            _ws = wb[_sn]
            if _ws.max_row == 1 and _ws.max_column == 1 and _ws["A1"].value is None:
                del wb[_sn]
        _write_student_sheet(wb, student_name, file_name, results, scoring, uncovered)
        wb.save(xlsx_path)
    else:  # pragma: no cover
        # 无 openpyxl 时兜底：仍生成 HTML，xlsx 不可用
        xlsx_path = None

    # ---- HTML（保留，与 xlsx 明细同构） ----
    reports = out / "reports"
    reports.mkdir(parents=True, exist_ok=True)
    pct = scoring["pct"]
    level = scoring.get("level") or level_of(pct)

    def row(r):
        c = r["criterion"]
        cls = "pass" if r["ok"] else ("review" if r.get("need_review") else "fail")
        verdict = "通过" if r["ok"] else ("待复核" if r.get("need_review") else "未通过")
        pos = r.get("position") or _position_from_scope(c.get("scope"))
        actual = _fmt(r.get("actual"))
        expected = _fmt(r.get("expected"))
        deduct = c.get("deduct", "") or ("" if r["ok"] else "未达到要求")
        return (
            f"<tr class='{cls}'><td>{c.get('label','')}</td><td>{verdict}</td>"
            f"<td>{r['weight']}</td><td>{actual}</td><td>{expected}</td>"
            f"<td>{pos}</td><td>{r.get('evidence','')}</td><td>{deduct}</td></tr>"
        )

    items = "".join(row(r) for r in results)
    uncovered_html = ""
    if uncovered:
        uncovered_html = "<h2>未覆盖需求</h2><ul>%s</ul>" % "".join(
            "<li>%s</li>" % u for u in uncovered)
    html = f"""<!DOCTYPE html><html lang="zh"><head><meta charset="utf-8"><title>作业评分报告</title>
<style>
body{{font-family:'Microsoft YaHei',sans-serif;margin:24px}}
table{{border-collapse:collapse;width:100%;table-layout:fixed}}
th,td{{border:1px solid #ccc;padding:6px 10px;font-size:13px;text-align:left;word-break:break-all}}
th{{background:#f5f5f5}}
.pass{{background:#e8f5e9}}.fail{{background:#ffebee}}.review{{background:#fff8e1}}
h1{{font-size:20px}}.score{{font-size:26px;color:#1565c0}}
</style></head><body><h1>作业评分报告</h1>
<p>文件：{file_name}　学生：{student_name}</p>
<p class="score">总分 {scoring['score']} / {scoring['max_score']}（{pct}%）　等级：{level}</p>
<h2>逐条明细</h2>
<table><tr><th style="width:14%">评分项</th><th style="width:6%">结论</th><th style="width:6%">分值</th>
<th style="width:12%">实际值</th><th style="width:12%">期望值</th><th style="width:12%">位置</th>
<th>依据</th><th style="width:10%">扣分说明</th></tr>{items}</table>
{uncovered_html}
</body></html>"""
    (reports / ("%s_report.html" % student_name)).write_text(html, encoding="utf-8")

    return {"summary_xlsx": str(xlsx_path) if xlsx_path else None,
            "report_html": str(reports / ("%s_report.html" % student_name))}
